# -*- coding: utf-8 -*-
"""
Created on Wed Jul 31 17:33:13 2019

@author: GuoOu
"""

import openpyxl

fileName = 'japan_tour.xlsx'
#load excel file
wb = openpyxl.load_workbook(fileName)

for sheet_name in wb.sheetnames:
    #get sheet
    sheet = wb[sheet_name]
    #process sheet
    for row in range(sheet.max_row):
        line = ''
        for col in range(sheet.max_column):
            cell = sheet.cell(row + 1, col + 1)
            line += str(cell.value) + ' '
        print(line)

wb.close()

