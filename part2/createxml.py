# -*- coding: utf-8 -*-
"""
Created on Wed Jul 31 18:00:33 2019

@author: GuoOu
"""
import openpyxl

wb = openpyxl.Workbook()
sheet_name = wb.sheetnames
sheet = wb.active
sheet.title = 'New working sheet'
#wb.create_sheet(index=1, title='Second Sheet')

#do your stuff
for row in range(10):
    for col in range(3):
        cell = sheet.cell(row + 1, col + 1)
        cell.value = row + col

max_row = sheet.max_row
for col in range(sheet.max_column):
    cell = sheet.cell(max_row + 1, col + 1)
    cell.value = '=sum(' + cell.column_letter + str(1) + ':' + cell.column_letter + str(max_row - 1) + ')'

#set row or column dimentions
sheet.row_dimensions[1].height = 20
sheet.column_dimensions['A'].width = 15

#freeze row 1
sheet.freeze_panes = 'A2'

wb.save('example.xlsx')
wb.close()